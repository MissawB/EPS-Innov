import glob
import subprocess  # Pour ouvrir le fichier sous Linux / macOS
import openpyxl
from kivy.uix.screenmanager import Screen
from kivy.uix.scrollview import ScrollView
from datetime import datetime
from kivy.uix.spinner import Spinner
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.uix.popup import Popup
from kivy.uix.label import Label
from kivy.uix.gridlayout import GridLayout
from kivy.utils import platform
from pathlib import Path

class ArchivesScreen(Screen):
    def __init__(self, **kwargs):
        super(ArchivesScreen, self).__init__(**kwargs)

        self.current_mode = "simplified"  # Mode par défaut : "simplified"

        # Layout principal
        main_layout = BoxLayout(orientation='vertical')

        # Layout pour les options de filtre, tri et mode
        top_layout = BoxLayout(orientation='horizontal', size_hint_y=None, height=50)

        # Bouton de retour
        self.back_button = Button(text='Retour', size_hint_x=0.2, height=50)
        self.back_button.bind(on_release=self.go_back)
        top_layout.add_widget(self.back_button)

        # Spinner pour sélectionner le filtre par sport
        self.filter_spinner = Spinner(
            text='Tout',
            values=['Tout', 'Futsal', 'Duel', 'Basket', 'Personnalisé'],
            size_hint_x=0.3
        )
        self.filter_spinner.bind(text=self.load_archives)
        top_layout.add_widget(self.filter_spinner)

        # Spinner pour le critère de tri
        self.sort_spinner = Spinner(
            text='Trier par',
            values=['Sport', 'Joueur', 'Variable', 'Nom d\'archive', 'Date'],
            size_hint_x=0.3
        )
        self.sort_spinner.bind(text=self.update_archive_display)
        top_layout.add_widget(self.sort_spinner)

        # Bouton pour changer de mode d'affichage
        self.mode_button = Button(text="Mode simplifié", size_hint_x=0.2, height=50)
        self.mode_button.bind(on_release=self.toggle_mode)
        top_layout.add_widget(self.mode_button)

        main_layout.add_widget(top_layout)

        # Layout pour afficher les archives avec scroll
        self.archive_layout = GridLayout(cols=1, size_hint_y=None, padding=10, spacing=10)
        self.archive_layout.bind(minimum_height=self.archive_layout.setter('height'))

        scroll_view = ScrollView(size_hint=(1, 1))
        scroll_view.add_widget(self.archive_layout)

        main_layout.add_widget(scroll_view)
        self.add_widget(main_layout)

        # Charger toutes les archives en mode simplifié dès le début
        self.load_archives()
        self.update_archive_display_mode()

    def toggle_mode(self, instance):
        """Basculer entre les modes d'affichage (simplifié ou détaillé)."""
        self.current_mode = "simplified" if self.current_mode == "detailed" else "detailed"
        self.mode_button.text = "Mode détaillé" if self.current_mode == "simplified" else "Mode simplifié"
        self.update_archive_display()  # Mettre à jour l'affichage selon le mode et le tri

    def update_archive_display_mode(self, *args) :
        """Met à jour l'affichage des archives en fonction du mode."""
        self.archive_layout.clear_widgets()
        if self.current_mode == "detailed":
            print("Affichage du mode détaillé")  # Debugging
            self.display_detailed_mode()
        elif self.current_mode == "simplified":
            print("Affichage du mode simplifié")  # Debugging
            self.display_simplified_mode()

    def display_detailed_mode(self):
        """Afficher les archives avec leurs détails."""
        # Logique pour le mode détaillé (déjà implémentée)
        self.display_default_order()

    def display_simplified_mode(self):
        """Afficher uniquement les noms des archives avec des boutons en mode simplifié, avec filtre."""
        script_path = Path(__file__).parent  # Répertoire du script
        archive_path = script_path / 'archives'  # Chemin complet du dossier "archives"

        # Lister les fichiers .txt dans le dossier "archives"
        file_list = glob.glob(str(archive_path / '*.txt'))  # Récupère la liste des fichiers .txt

        if not file_list:
            print("Aucune archive trouvée dans le dossier 'archives'.")
            return

        # Liste pour stocker les archives avec leurs informations (date, sport)
        archives_with_info = []

        selected_sport = self.filter_spinner.text.strip().lower()
        print(f"Filtre appliqué : {selected_sport}")

        for file_path in file_list:
            filename = Path(file_path).name  # Récupère le nom du fichier (ex : "match1.txt")

            # Lecture du contenu du fichier
            content = self.read_file_content(file_path)  # Assurez-vous que cette méthode lit bien le fichier
            if not content:
                print(f"Contenu vide pour {file_path}")
                continue

            # Extraire la date et le type de sport du contenu
            archive_date = self.extract_date(content)
            sport_type = self.extract_sport_type(content).lower()

            # Appliquer le filtre de sport
            if selected_sport != 'tout' and sport_type != selected_sport:
                continue  # Si le sport ne correspond pas au filtre, on passe à l'archive suivante

            if archive_date:
                archives_with_info.append((filename, file_path, archive_date, sport_type))

        # Trier les archives par date ou par nom
        sort_by = self.sort_spinner.text
        if sort_by == "Date":
            sorted_archives = sorted(archives_with_info, key=lambda x: x[2], reverse=True)  # Tri par date
        else:
            sorted_archives = sorted(archives_with_info, key=lambda x: x[0])  # Tri par nom d'archive

        # Affichage des archives triées
        self.archive_layout.clear_widgets()

        for filename, file_path, _, sport_type in sorted_archives:
            archive_layout = BoxLayout(orientation='horizontal', size_hint_y=None, height=40)

            # Nom du fichier sans l'extension
            name_without_extension = filename.split('.')[0]  # Enlève l'extension .txt
            archive_label = Label(text=name_without_extension, font_size=25, size_hint_x=0.7)
            archive_layout.add_widget(archive_label)

            # Affichage du sport
            sport_label = Label(text=f"Sport: {sport_type.capitalize()}", font_size=15, size_hint_x=0.2)
            archive_layout.add_widget(sport_label)

            # Bouton d'accès à l'archive
            access_button = Button(text="Ouvrir", size_hint_x=0.15)
            access_button.bind(on_release=lambda instance, path=file_path: self.open_archive_popup(path))
            archive_layout.add_widget(access_button)

            self.archive_layout.add_widget(archive_layout)

    def open_archive_popup(self, file_path):
        """Ouvrir une popup pour gérer l'archive."""

        # Lire le contenu de l'archive
        content = self.read_file_content(file_path)
        if not content:
            print(f"Le fichier {file_path} est vide ou illisible.")
            return

        # **Extraire la date de l'archive**
        date_archive = "Inconnue"  # Valeur par défaut si aucune date n'est trouvée
        for line in content:
            if line.startswith("Date et heure :"):
                date_archive = line.replace("Date et heure :", "").strip()
                break

        # Extraire le type de sport à partir du contenu de l'archive
        sport_type = self.extract_sport_type(content)

        # Créer la disposition de la popup
        popup_layout = BoxLayout(orientation='vertical', padding=10, spacing=10)

        # Ajouter un label pour afficher le sport en titre
        sport_label = Label(text=f"Sport: {sport_type}", bold=True, font_size=20, size_hint_y=None, height=40)
        popup_layout.add_widget(sport_label)

        # Créer un tableau pour afficher les informations de l'archive
        table_layout = GridLayout(cols=3, size_hint_y=None, padding=5, spacing=5)
        table_layout.bind(minimum_height=table_layout.setter('height'))

        # Ajouter les en-têtes du tableau : Joueur, Variable, Score
        table_layout.add_widget(Label(text="Joueur", bold=True, size_hint_y=None, height=40))
        table_layout.add_widget(Label(text="Variable", bold=True, size_hint_y=None, height=40))
        table_layout.add_widget(Label(text="Score", bold=True, size_hint_y=None, height=40))

        # Organiser les données par joueur et variable
        player_data = {}
        current_player = None

        for line in content:
            line = line.strip()
            if line.startswith("¤"):  # Début d'un joueur
                current_player = line[1:]
                if current_player not in player_data:
                    player_data[current_player] = []
            elif line.startswith("£") and current_player:  # Variables du joueur
                parts = line.split(":")
                if len(parts) == 2:
                    variable = parts[0].strip(" £")
                    score = parts[1].strip()
                    player_data[current_player].append((variable, score))

        # Ajouter les données des joueurs et variables au tableau
        for player, variables in player_data.items():
            for variable, score in variables:
                table_layout.add_widget(Label(text=player, size_hint_y=None, height=30))
                table_layout.add_widget(Label(text=variable, size_hint_y=None, height=30))  # Afficher la variable
                table_layout.add_widget(Label(text=score, size_hint_y=None, height=30))  # Afficher le score

        # Envelopper le tableau dans un ScrollView pour permettre le défilement
        scroll_view = ScrollView(size_hint=(1, 1))
        scroll_view.add_widget(table_layout)

        # Ajouter le ScrollView au layout principal de la popup
        popup_layout.add_widget(scroll_view)

        # **1. Créer d'abord la popup**
        archive_path = Path(file_path)
        archive_name = archive_path.stem  # Nom du fichier sans l'extension
        popup_title = f"Données de l'archive - {archive_name} ({date_archive})"  # Ajout de la date
        popup = Popup(title=popup_title, content=popup_layout, size_hint=(0.9, 0.9))

        # Ajouter les boutons de fermeture, de suppression et d'exportation
        button_layout = BoxLayout(orientation='horizontal', size_hint_y=None, height=50)

        # **2. Assurez-vous que close_button utilise la méthode dismiss() de la popup**
        close_button = Button(text="Fermer", size_hint_x=0.4)
        delete_button = Button(text="Supprimer", size_hint_x=0.3)
        export_button = Button(text="Exporter", size_hint_x=0.3)

        # **3. Lier la fonction de fermeture à la popup**
        close_button.bind(on_release=lambda instance: popup.dismiss())

        # Boutons de suppression et d'export
        delete_button.bind(
            on_release=lambda instance, path=str(archive_path), widget_to_remove=popup_layout:
            self.delete_archive(path, widget_to_remove)
        )
        export_button.bind(on_release=lambda instance: self.export_to_excel(file_path, player_data))

        # Ajouter les boutons au layout
        button_layout.add_widget(close_button)
        button_layout.add_widget(delete_button)
        button_layout.add_widget(export_button)

        popup_layout.add_widget(button_layout)

        # **4. Ouvrir la popup**
        try:
            popup.open()
        except Exception as e:
            print(f"❌ Erreur lors de l'ouverture de la popup : {e}")

    def export_to_excel(self, file_path, player_data):
        """Exporter les données de l'archive au format Excel."""

        # Détecter le bon chemin de base (Windows, Android, iOS)
        if platform == 'android':
            from android.storage import app_storage_path
            base_path = Path(app_storage_path())
        elif platform == 'ios':
            base_path = Path().home() / 'Documents'
        else:
            base_path = Path(__file__).parent

        # Nom du dossier d'exportation (il sera créé dans le bon répertoire)
        export_folder = base_path / "archives_excel"
        export_folder.mkdir(parents=True, exist_ok=True)  # Crée le dossier s'il n'existe pas

        # Nom du fichier Excel basé sur le nom du fichier de l'archive (sans .txt)
        base_filename = Path(file_path).stem  # Nom de fichier sans extension
        excel_path = export_folder / f"{base_filename}.xlsx"

        try:
            # Créer le fichier Excel
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Données de l'archive"

            # Ajouter les en-têtes au fichier Excel
            headers = ["Joueur", "Variable", "Score"]
            for col_num, header in enumerate(headers, start=1):
                sheet.cell(row=1, column=col_num, value=header)

            # Ajouter les données des joueurs au fichier Excel
            row_num = 2
            for player, variables in player_data.items():
                for variable, score in variables:
                    sheet.cell(row=row_num, column=1, value=player)  # Colonne Joueur
                    sheet.cell(row=row_num, column=2, value=variable)  # Colonne Variable
                    sheet.cell(row=row_num, column=3, value=score)  # Colonne Score
                    row_num += 1

            # Enregistrer le fichier Excel
            workbook.save(excel_path)

            # Ouvrir le fichier Excel après exportation
            if platform == 'win':  # Windows
                subprocess.Popen(["start", excel_path], shell=True)
            elif platform == 'macosx':  # macOS (iOS est basé sur Darwin mais n'a pas de Popen)
                subprocess.Popen(["open", excel_path])
            elif platform == 'linux':
                subprocess.Popen(["xdg-open", excel_path])
            elif platform == 'android':
                try:
                    from jnius import autoclass
                    Intent = autoclass('android.content.Intent')
                    Uri = autoclass('android.net.Uri')
                    File = autoclass('java.io.File')
                    PythonActivity = autoclass('org.kivy.android.PythonActivity')

                    file = File(str(excel_path))
                    uri = Uri.fromFile(file)
                    intent = Intent()
                    intent.setAction(Intent.ACTION_VIEW)
                    intent.setDataAndType(uri, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    currentActivity = PythonActivity.mActivity
                    currentActivity.startActivity(intent)
                except Exception as e:
                    print(f"Erreur lors de l'ouverture du fichier sur Android : {e}")
            else:
                print(f"Plateforme non prise en charge pour l'ouverture automatique du fichier Excel.")

            # Afficher une popup de confirmation (si nécessaire)
            self.show_export_confirmation_popup(str(excel_path))

        except Exception as e:
            print(f"Erreur lors de l'exportation du fichier Excel : {e}")

    def show_export_confirmation_popup(self, excel_path):
        """Afficher une popup de confirmation d'exportation."""
        layout = BoxLayout(orientation='vertical', padding=10, spacing=10)
        message = f"Le fichier a été exporté avec succès :\n{excel_path}"
        label = Label(text=message, halign="center", valign="middle")
        label.text_size = (400, None)  # Limite la largeur du texte
        button = Button(text="OK", size_hint_y=None, height=50)
        popup = Popup(title="Exportation réussie", content=layout, size_hint=(0.8, 0.4))
        button.bind(on_release=popup.dismiss)
        layout.add_widget(label)
        layout.add_widget(button)
        popup.open()

    def load_archives(self, *args):
        """Charger et afficher les archives, en appliquant le filtre sport."""
        self.archive_layout.clear_widgets()
        self.sorted_data = {"Sport": {}, "Joueur": {}, "Variable": {}}

        # Détecter le bon chemin de base (Windows, Android, iOS, etc.)
        if platform == 'android':
            from android.storage import app_storage_path
            base_path = Path(app_storage_path())
        elif platform == 'ios':
            base_path = Path().home() / 'Documents'
        else:
            base_path = Path(__file__).parent

        # Chemin du dossier des archives
        archive_dir = base_path / "archives"

        if not archive_dir.exists():
            print("Le dossier des archives n'existe pas.")
            return

        selected_sport = self.filter_spinner.text.strip().lower()
        print(f"Chargement des archives avec le filtre sport : {selected_sport}")

        archives_with_dates = []

        # Parcourir tous les fichiers du dossier des archives
        for file_path in sorted(archive_dir.iterdir(), reverse=True):
            if file_path.is_file() and file_path.suffix == '.txt':  # On ne considère que les fichiers .txt
                content = self.read_file_content(file_path)

                if content is None:
                    continue

                sport_type = self.extract_sport_type(content).lower()
                if selected_sport != 'tout' and sport_type != selected_sport:
                    continue

                archive_date = self.extract_date(content)
                if archive_date:
                    archives_with_dates.append((file_path.name, file_path, content, sport_type, archive_date))
                    self.organize_data(file_path.name, content, sport_type)

        # Trier les archives par date décroissante
        sorted_archives_by_date = sorted(archives_with_dates, key=lambda x: x[4], reverse=True)

        # Affichage des archives triées
        for filename, file_path, content, sport_type, _ in sorted_archives_by_date:
            self.add_archive_to_layout(filename, content, str(file_path), sport_type)

    def is_detailed_mode(self):
        """Retourne True si le mode détaillé est activé, sinon False (mode simplifié)."""
        return self.mode_spinner.text.strip().lower() == 'détaillé'

    def organize_data(self, filename, content, sport_type):
        player_data = {}
        variable_data = {}  # Nouvelle structure pour organiser les données par variable
        current_player = None
        for line in content:
            line = line.strip()
            if line.startswith("¤"):
                current_player = line[1:]
                if current_player not in player_data:
                    player_data[current_player] = {'variables': []}
            elif line.startswith("£") and current_player:
                parts = line.split(":")
                variable = parts[0].strip(" £")
                score = int(parts[1].strip())
                player_data[current_player]['variables'].append((variable, score))

                # Ajouter la variable aux données globales
                if variable not in variable_data:
                    variable_data[variable] = []
                variable_data[variable].append((current_player, score))

        for player, data in player_data.items():
            if sport_type not in self.sorted_data["Sport"]:
                self.sorted_data["Sport"][sport_type] = []
            self.sorted_data["Sport"][sport_type].append((player, data))

            if player not in self.sorted_data["Joueur"]:
                self.sorted_data["Joueur"][player] = {}
            self.sorted_data["Joueur"][player][sport_type] = data

        # Ajouter les données triées par variable dans le dictionnaire global
        self.sorted_data["Variable"] = variable_data

        # Lecture du contenu des fichiers

    @staticmethod
    def read_file_content(file_path):
        """Tente de lire le fichier en utilisant plusieurs encodages."""

        # Assurez-vous que file_path est bien de type Path
        if not isinstance(file_path, Path):
            file_path = Path(file_path)

        # Liste des encodages à essayer
        encodings = ['iso-8859-1', 'utf-8', 'latin-1']  # Attention : il manquait une virgule entre 'utf-8' et 'latin-1'

        for enc in encodings:
            try:
                with file_path.open('r', encoding=enc) as f:  # Utilisation de Path().open() au lieu de open()
                    return f.readlines()
            except UnicodeDecodeError:
                print(f"Erreur de décodage avec l'encodage {enc} pour le fichier {file_path}")
            except Exception as e:
                print(f"Erreur inconnue lors de la lecture du fichier {file_path} : {e}")

        print(f"Impossible de lire le fichier {file_path} avec les encodages disponibles.")
        return None

    @staticmethod
    def extract_sport_type(content):
        for line in content:
            if "Personnalisé" in line:
                return "Personnalisé"
            if "Futsal" in line:
                return "Futsal"
            if "Duel" in line:
                return "Duel"
            if "Basket" in line:
                return "Basket"
        return "Inconnu"

    @staticmethod
    def extract_date(content):
        try:
            # Vérifier si la ligne existe
            if len(content) > 1:
                date_line = content[1]
                # Chercher le premier ':' et extraire tout ce qui suit
                if ":" in date_line:
                    date_str = date_line.split(":", 1)[1].strip()  # Séparer uniquement à la première occurrence de ":"
                    # Vérifier si la date est bien dans le bon format
                    return datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
                else:
                    print("Format de la ligne de date invalide : pas de ':' trouvé.")
            else:
                print("La ligne de date est manquante dans le contenu.")
        except Exception as e:
            print(f"Erreur lors de l'extraction de la date : {e}")
        return None

    def update_archive_display(self, *args):
        """Met à jour l'affichage des archives en fonction du tri et du mode (simplifié ou détaillé)."""
        self.archive_layout.clear_widgets()
        sort_by = self.sort_spinner.text

        if self.current_mode == "simplified":
            self.display_simplified_mode()
        else:
            if sort_by == "Nom d'archive":
                self.display_sorted_by_filename()
            elif sort_by == "Date":
                self.display_sorted_by_date()
            elif sort_by == "Sport":
                self.display_sorted_data('Sport')
            elif sort_by == "Joueur":
                self.display_sorted_data('Joueur')
            elif sort_by == "Variable":
                self.display_sorted_by_variable()
            else:
                self.display_default_order()

    def add_archive_to_layout(self, filename, content, file_path, sport_type):
        # Vérifier si l'archive doit être affichée sous forme de tableau
        selected_sport = self.filter_spinner.text.lower()

        if selected_sport == 'tout' or selected_sport == sport_type:
            title_layout = BoxLayout(orientation='horizontal', size_hint_y=None, height=40)
            archive_label = Label(text=f"Archive: {filename} - Sport: {sport_type.capitalize()}", font_size=18,
                                  bold=True)
            title_layout.add_widget(archive_label)

            delete_button = Button(text="Effacer", size_hint_x=0.2)
            delete_button.bind(on_release=lambda instance: self.delete_archive(file_path, title_layout))
            title_layout.add_widget(delete_button)

            self.archive_layout.add_widget(title_layout)

            # Créer le tableau pour afficher les données de l'archive
            table_layout = GridLayout(cols=3, size_hint_y=None, padding=5, spacing=5)
            table_layout.bind(minimum_height=table_layout.setter('height'))
            table_layout.add_widget(Label(text="Joueur", bold=True, size_hint_y=None, height=30))
            table_layout.add_widget(Label(text="Variable", bold=True, size_hint_y=None, height=30))
            table_layout.add_widget(Label(text="Score", bold=True, size_hint_y=None, height=30))

            # Extraction des données de chaque joueur et variable
            player_data = {}
            current_player = None

            for line in content:
                line = line.strip()
                if line.startswith("¤"):
                    current_player = line[1:]
                    player_data[current_player] = {'variables': []}
                elif line.startswith("£") and current_player:
                    parts = line.split(":")
                    variable = parts[0].strip(" £")
                    score = int(parts[1].strip())
                    player_data[current_player]['variables'].append((variable, score))

            # Ajout de toutes les variables pour chaque joueur
            for player, data in player_data.items():
                for variable, score in data['variables']:
                    table_layout.add_widget(Label(text=player, size_hint_y=None, height=30))
                    table_layout.add_widget(Label(text=variable, size_hint_y=None, height=30))
                    table_layout.add_widget(Label(text=str(score), size_hint_y=None, height=30))

            self.archive_layout.add_widget(table_layout)

    def display_default_order(self):
        """Afficher les archives sans tri (ordre par défaut)."""

        # Définition du dossier des archives
        archive_dir = Path("archives")  # Utilisation de Path au lieu de chaîne de caractères

        if not archive_dir.exists():
            print(f"Le dossier {archive_dir} n'existe pas.")
            return

        # Tri des fichiers par nom dans l'ordre décroissant
        for file_path in sorted(archive_dir.iterdir(), key=lambda x: x.name, reverse=True):
            if file_path.is_file():  # Vérifie que c'est bien un fichier
                content = self.read_file_content(file_path)
                if content:
                    sport_type = self.extract_sport_type(content)
                    self.add_archive_to_layout(file_path.name, content, file_path, sport_type)

    def display_sorted_by_filename(self):
        """Trier les archives par nom d'archive (ordre alphabétique) et les afficher."""

        # Chemin du dossier des archives
        archive_dir = Path("archives")  # Utilisation de Path au lieu d'une simple chaîne de caractères

        if not archive_dir.exists():
            print(f"Le dossier {archive_dir} n'existe pas.")
            return

        # Tri alphabétique des fichiers
        sorted_files = sorted(archive_dir.iterdir(), key=lambda x: x.name)  # Tri par nom de fichier

        for file_path in sorted_files:
            if file_path.is_file():  # Vérifie qu'il s'agit bien d'un fichier
                content = self.read_file_content(file_path)
                if content:
                    sport_type = self.extract_sport_type(content)
                    self.add_archive_to_layout(file_path.name, content, file_path, sport_type)

    def display_sorted_by_date(self):
        """Trier les archives par date et afficher les détails de chaque archive."""

        # Chemin du dossier des archives
        archive_dir = Path("archives")
        archives_with_dates = []

        if not archive_dir.exists():
            print(f"❌ Le dossier {archive_dir} n'existe pas.")
            return

        # Parcourir tous les fichiers du dossier archives
        for file_path in archive_dir.iterdir():
            try:
                if file_path.is_file():  # Vérifie qu'il s'agit bien d'un fichier
                    content = self.read_file_content(file_path)

                    if not content:
                        print(f"⚠️ Fichier vide ou contenu non lisible : {file_path.name}")
                        continue

                    # Extraction de la date à partir du contenu
                    archive_date = self.extract_date(content)
                    if archive_date:
                        archives_with_dates.append((file_path.name, file_path, content, archive_date))
                    else:
                        print(f"⚠️ Date non trouvée ou invalide dans le fichier {file_path.name}")

            except Exception as e:
                print(f"❌ Erreur de lecture ou d'extraction des données de {file_path.name} : {e}")

        # Trier les archives par date (du plus récent au plus ancien)
        try:
            sorted_archives_by_date = sorted(archives_with_dates, key=lambda x: x[3], reverse=True)
        except Exception as e:
            print(f"❌ Erreur de tri des dates : {e}")
            sorted_archives_by_date = []

        # Ajouter les archives triées à l'interface
        for filename, file_path, content, archive_date in sorted_archives_by_date:
            try:
                # Formater la date pour l'affichage (format : "YYYY-MM-DD HH:MM:SS")
                formatted_date = archive_date.strftime("%Y-%m-%d %H:%M:%S")

                # Ajouter l'archive à l'interface
                self.add_archive_to_layout(filename, content, file_path, formatted_date)
            except Exception as e:
                print(f"❌ Erreur lors de l'ajout de l'archive {filename} à l'interface : {e}")

    def display_sorted_by_variable(self):
        # Créer un dictionnaire pour stocker les variables avec leurs scores respectifs pour chaque joueur
        variable_data = {}

        # Parcourir toutes les archives pour extraire les variables et les scores
        for sport, records in self.sorted_data["Sport"].items():
            for player, data in records:
                for variable, score in data['variables']:
                    # Si la variable n'existe pas dans le dictionnaire, l'ajouter
                    if variable not in variable_data:
                        variable_data[variable] = []
                    # Ajouter le score du joueur à la variable correspondante
                    variable_data[variable].append((player, sport, score))

        # Trier les variables par nom (ordre alphabétique)
        sorted_variables = sorted(variable_data.keys())

        # Afficher les variables triées
        for variable in sorted_variables:
            title_label = Label(text=f"Variable: {variable}", font_size=18, bold=True, size_hint_y=None, height=30)
            self.archive_layout.add_widget(title_label)

            table_layout = GridLayout(cols=3, size_hint_y=None, padding=5, spacing=5)
            table_layout.bind(minimum_height=table_layout.setter('height'))
            table_layout.add_widget(Label(text="Joueur", bold=True, size_hint_y=None, height=30))
            table_layout.add_widget(Label(text="Sport", bold=True, size_hint_y=None, height=30))
            table_layout.add_widget(Label(text="Score", bold=True, size_hint_y=None, height=30))

            # Afficher les scores des joueurs pour cette variable
            for player, sport, score in variable_data[variable]:
                table_layout.add_widget(Label(text=player, size_hint_y=None, height=30))
                table_layout.add_widget(Label(text=sport.capitalize(), size_hint_y=None, height=30))
                table_layout.add_widget(Label(text=str(score), size_hint_y=None, height=30))

            self.archive_layout.add_widget(table_layout)

    def display_sorted_data(self, sort_by):
        if sort_by == "Joueur":
            for player, sports in self.sorted_data["Joueur"].items():
                title_label = Label(text=f"Joueur: {player}", font_size=18, bold=True, size_hint_y=None, height=30)
                self.archive_layout.add_widget(title_label)

                table_layout = GridLayout(cols=3, size_hint_y=None, padding=5, spacing=5)
                table_layout.bind(minimum_height=table_layout.setter('height'))
                table_layout.add_widget(Label(text="Sport", bold=True, size_hint_y=None, height=30))
                table_layout.add_widget(Label(text="Variable", bold=True, size_hint_y=None, height=30))
                table_layout.add_widget(Label(text="Score", bold=True, size_hint_y=None, height=30))

                for sport, data in sports.items():
                    for variable, score in data['variables']:
                        table_layout.add_widget(Label(text=sport.capitalize(), size_hint_y=None, height=30))
                        table_layout.add_widget(Label(text=variable, size_hint_y=None, height=30))
                        table_layout.add_widget(Label(text=str(score), size_hint_y=None, height=30))

                self.archive_layout.add_widget(table_layout)

        elif sort_by == "Sport":
            for sport, records in self.sorted_data["Sport"].items():
                title_label = Label(text=f"Sport: {sport}", font_size=18, bold=True, size_hint_y=None, height=30)
                self.archive_layout.add_widget(title_label)

                table_layout = GridLayout(cols=3, size_hint_y=None, padding=5, spacing=5)
                table_layout.bind(minimum_height=table_layout.setter('height'))
                table_layout.add_widget(Label(text="Joueur", bold=True, size_hint_y=None, height=30))
                table_layout.add_widget(Label(text="Variable", bold=True, size_hint_y=None, height=30))
                table_layout.add_widget(Label(text="Score", bold=True, size_hint_y=None, height=30))

                for player, data in records:
                    for variable, score in data['variables']:
                        table_layout.add_widget(Label(text=player, size_hint_y=None, height=30))
                        table_layout.add_widget(Label(text=variable, size_hint_y=None, height=30))
                        table_layout.add_widget(Label(text=str(score), size_hint_y=None, height=30))

                self.archive_layout.add_widget(table_layout)
        else:
            # Autres critères de tri (par exemple Variable)
            for key, records in self.sorted_data["Joueur"].items():
                title_label = Label(text=f"{sort_by}: {key}", font_size=18, bold=True, size_hint_y=None, height=30)
                self.archive_layout.add_widget(title_label)

                table_layout = GridLayout(cols=3, size_hint_y=None, padding=5, spacing=5)
                table_layout.bind(minimum_height=table_layout.setter('height'))
                table_layout.add_widget(
                    Label(text="Sport", bold=True, size_hint_y=None, height=30))  # Remplacer "Joueur" par "Sport"
                table_layout.add_widget(Label(text="Variable", bold=True, size_hint_y=None, height=30))
                table_layout.add_widget(Label(text="Score", bold=True, size_hint_y=None, height=30))

                # Affichage des données triées par Sport, Variable, etc.
                for player, data in records:
                    for variable, score in data['variables']:
                        # Afficher le sport de l'archive
                        sport_type = self.get_sport_for_player(player)  # Récupérer le sport associé à l'archive
                        table_layout.add_widget(
                            Label(text=sport_type.capitalize(), size_hint_y=None, height=30))  # Afficher le sport
                        table_layout.add_widget(Label(text=variable, size_hint_y=None, height=30))
                        table_layout.add_widget(Label(text=str(score), size_hint_y=None, height=30))

                self.archive_layout.add_widget(table_layout)

    def get_sport_for_player(self, player_name):
        """ Trouver le sport pour un joueur donné """
        for sport, players in self.sorted_data["Sport"].items():
            for player, _ in players:
                if player == player_name:
                    return sport
        return "Inconnu"

    def open_archive(self, file_path):
        """Ouvrir le fichier d'archive et afficher son contenu."""

        # Convertir le chemin en objet Path
        archive_path = Path(file_path)

        # Vérifier si le fichier existe et est un fichier
        if not archive_path.is_file():
            print(f"❌ Le fichier {archive_path} n'existe pas ou n'est pas un fichier valide.")
            return

        try:
            # Lire le contenu du fichier
            content = self.read_file_content(archive_path)
            if content:
                print("\n".join(content))
            else:
                print(f"⚠️ Le fichier {archive_path.name} est vide ou le contenu n'a pas pu être lu.")
        except Exception as e:
            print(f"❌ Erreur lors de l'ouverture du fichier {archive_path.name} : {e}")

    #Bouton effacer
    def delete_archive(self, file_path, widget_to_remove):
        """Supprimer une archive avec confirmation de l'utilisateur."""

        def confirm_deletion(instance):
            try:
                archive_path = Path(file_path)

                # Vérifier si le fichier existe avant de le supprimer
                if archive_path.is_file():
                    archive_path.unlink()  # Supprime le fichier
                    self.archive_layout.remove_widget(widget_to_remove)
                    print(f"✅ Archive supprimée : {archive_path.name}")
                else:
                    print(f"❌ Le fichier {archive_path} n'existe pas ou n'est pas un fichier valide.")

                popup.dismiss()
            except Exception as e:
                print(f"❌ Erreur lors de la suppression de l'archive {archive_path.name} : {e}")

        # Création de la boîte de confirmation
        popup_layout = BoxLayout(orientation='vertical', padding=10, spacing=10)
        label = Label(text="Êtes-vous sûr de vouloir supprimer cette archive ?")

        # Boutons de confirmation et d'annulation
        buttons_layout = BoxLayout(orientation='horizontal', size_hint_y=None, height=50)
        confirm_button = Button(text="Oui", on_release=confirm_deletion)
        cancel_button = Button(text="Non", on_release=lambda x: popup.dismiss())
        buttons_layout.add_widget(confirm_button)
        buttons_layout.add_widget(cancel_button)

        # Ajout des widgets à la mise en page de la popup
        popup_layout.add_widget(label)
        popup_layout.add_widget(buttons_layout)

        # Création et ouverture de la popup
        popup = Popup(title="Confirmation de suppression", content=popup_layout, size_hint=(0.8, 0.4))
        popup.open()

    #Bouton retour
    def go_back(self, instance):
        self.manager.current = 'home'